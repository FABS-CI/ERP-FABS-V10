"""
Module Rapports — EDITIONS FABS-CI ERP
- Rapports de ventes avec filtres multiples
- Rapports de stock avec alertes
- Exports PDF
"""
from fastapi import APIRouter, HTTPException, Request, Header
from typing import Optional, List, Dict, Any
from datetime import datetime, timezone
from collections import defaultdict
import logging

logger = logging.getLogger("fabsci")


def build_rapports_router(db, resolve_user):
    router = APIRouter(prefix="/rapports", tags=["Rapports"])
    
    # RBAC
    READ_ROLES = {"super_admin", "directeur_general", "comptable", "directeur_commercial"}
    
    async def check_read(request, authorization):
        u = await resolve_user(request, authorization)
        if u["role"] not in READ_ROLES:
            raise HTTPException(status_code=403, detail="Accès interdit")
        return u
    
    @router.get("/ventes")
    async def get_rapport_ventes(
        request: Request,
        authorization: Optional[str] = Header(default=None),
        matiere: Optional[str] = None,
        ecole: Optional[str] = None,
        localite: Optional[str] = None,
        niveau_scolaire: Optional[str] = None,
        date_debut: Optional[str] = None,
        date_fin: Optional[str] = None,
    ):
        """
        Rapport de ventes avec filtres multiples.
        Retourne les données agrégées par matière, école, localité, niveau.
        """
        await check_read(request, authorization)
        
        # Construction du pipeline d'agrégation MongoDB
        # Étape 1 : Récupérer toutes les factures émises/payées
        match_stage = {
            "statut": {"$in": ["emise", "partiellement_payee", "payee"]},
            "type_facture": "facture"
        }
        
        if date_debut:
            match_stage["date_facture"] = {"$gte": date_debut}
        if date_fin:
            if "date_facture" in match_stage:
                match_stage["date_facture"]["$lte"] = date_fin
            else:
                match_stage["date_facture"] = {"$lte": date_fin}
        
        factures = await db.factures.find(match_stage, {"_id": 0}).to_list(1000)
        
        # Étape 2 : Pour chaque facture, récupérer les lignes
        results = []
        total_quantite = 0
        total_montant = 0
        
        for facture in factures:
            facture_id = facture["facture_id"]
            client_id = facture.get("client_id")
            
            # Récupérer le client
            client = await db.clients.find_one({"client_id": client_id}, {"_id": 0})
            if not client:
                continue
            
            # Vérifier les filtres client
            if ecole and client.get("nom_client", "").lower().find(ecole.lower()) == -1:
                continue
            if localite and client.get("localite", "").lower().find(localite.lower()) == -1:
                continue
            
            # Récupérer les lignes de facture
            lignes = await db.facture_lignes.find({"facture_id": facture_id}, {"_id": 0}).to_list(100)
            
            for ligne in lignes:
                produit_id = ligne.get("produit_id")
                produit = await db.produits.find_one({"product_id": produit_id}, {"_id": 0})
                
                if not produit:
                    continue
                
                # Appliquer les filtres produit
                if matiere:
                    # La matière peut être dans le titre ou auteur
                    titre_auteur = f"{produit.get('titre', '')} {produit.get('auteur', '')}".lower()
                    if matiere.lower() not in titre_auteur:
                        continue
                
                if niveau_scolaire and produit.get("niveau_scolaire", "").lower().find(niveau_scolaire.lower()) == -1:
                    continue
                
                # Calculer les montants
                quantite = ligne.get("quantite", 0)
                montant_ht = ligne.get("montant_ht", 0)
                
                # Ajouter au résultat
                results.append({
                    "matiere": produit.get("auteur", "Non spécifié"),  # Utiliser auteur comme proxy pour matière
                    "titre_produit": produit.get("titre", ""),
                    "ecole": client.get("nom_client", ""),
                    "localite": client.get("localite", "Non spécifiée"),
                    "niveau_scolaire": produit.get("niveau_scolaire", "Non spécifié"),
                    "quantite_vendue": quantite,
                    "montant_total": montant_ht,
                    "date_facture": facture.get("date_facture", ""),
                    "reference_facture": facture.get("reference", ""),
                })
                
                total_quantite += quantite
                total_montant += montant_ht
        
        # Agrégations pour les graphiques
        # Par matière
        ventes_par_matiere = defaultdict(lambda: {"quantite": 0, "montant": 0})
        for r in results:
            matiere = r["matiere"]
            ventes_par_matiere[matiere]["quantite"] += r["quantite_vendue"]
            ventes_par_matiere[matiere]["montant"] += r["montant_total"]
        
        # Par localité
        ventes_par_localite = defaultdict(lambda: {"quantite": 0, "montant": 0})
        for r in results:
            loc = r["localite"]
            ventes_par_localite[loc]["quantite"] += r["quantite_vendue"]
            ventes_par_localite[loc]["montant"] += r["montant_total"]
        
        # Par date (agrégation par mois)
        ventes_par_mois = defaultdict(lambda: {"quantite": 0, "montant": 0})
        for r in results:
            if r["date_facture"]:
                mois = r["date_facture"][:7]  # YYYY-MM
                ventes_par_mois[mois]["quantite"] += r["quantite_vendue"]
                ventes_par_mois[mois]["montant"] += r["montant_total"]
        
        return {
            "lignes": results,
            "total_quantite": total_quantite,
            "total_montant": total_montant,
            "nombre_lignes": len(results),
            "agregations": {
                "par_matiere": [
                    {"matiere": k, **v} for k, v in ventes_par_matiere.items()
                ],
                "par_localite": [
                    {"localite": k, **v} for k, v in ventes_par_localite.items()
                ],
                "par_mois": [
                    {"mois": k, **v} for k, v in sorted(ventes_par_mois.items())
                ],
            }
        }
    
    @router.get("/stock")
    async def get_rapport_stock(
        request: Request,
        authorization: Optional[str] = Header(default=None),
        matiere: Optional[str] = None,
        niveau_scolaire: Optional[str] = None,
        alerte_uniquement: bool = False,
    ):
        """
        Rapport de stock avec alertes.
        """
        await check_read(request, authorization)
        
        # Récupérer tous les produits
        match_stage = {"actif": True}
        
        produits = await db.produits.find(match_stage, {"_id": 0}).to_list(1000)
        
        results = []
        total_stock_valeur = 0
        nb_alertes = 0
        
        for produit in produits:
            # Filtres
            if matiere:
                titre_auteur = f"{produit.get('titre', '')} {produit.get('auteur', '')}".lower()
                if matiere.lower() not in titre_auteur:
                    continue
            
            if niveau_scolaire and produit.get("niveau_scolaire", "").lower().find(niveau_scolaire.lower()) == -1:
                continue
            
            stock_actuel = produit.get("stock_actuel", 0)
            stock_minimum = produit.get("stock_minimum", 10)
            prix_vente = produit.get("prix_vente", 0)
            
            # Alerte si stock < minimum
            en_alerte = stock_actuel < stock_minimum
            
            if alerte_uniquement and not en_alerte:
                continue
            
            if en_alerte:
                nb_alertes += 1
            
            # Calculer la valeur du stock
            valeur_stock = stock_actuel * prix_vente
            total_stock_valeur += valeur_stock
            
            results.append({
                "reference": produit.get("reference", ""),
                "titre": produit.get("titre", ""),
                "auteur": produit.get("auteur", ""),
                "niveau_scolaire": produit.get("niveau_scolaire", ""),
                "categorie": produit.get("categorie", ""),
                "stock_actuel": stock_actuel,
                "stock_minimum": stock_minimum,
                "prix_vente": prix_vente,
                "valeur_stock": valeur_stock,
                "en_alerte": en_alerte,
                "statut_stock": "rupture" if stock_actuel == 0 else ("alerte" if en_alerte else "ok"),
            })
        
        # Récupérer l'historique des mouvements récents (30 derniers) avec $lookup
        pipeline_mouvements = [
            {"$sort": {"date_mouvement": -1}},
            {"$limit": 30},
            {"$lookup": {
                "from": "produits",
                "localField": "product_id",
                "foreignField": "product_id",
                "as": "produit_info"
            }},
            {"$addFields": {
                "titre_produit": {"$arrayElemAt": ["$produit_info.titre", 0]}
            }},
            {"$project": {
                "produit_info": 0,
                "_id": 0
            }}
        ]
        mouvements = await db.mouvements_stock.aggregate(pipeline_mouvements).to_list(30)
        
        return {
            "produits": results,
            "total_produits": len(results),
            "nb_alertes": nb_alertes,
            "valeur_stock_total": total_stock_valeur,
            "mouvements_recents": mouvements,
        }
    
    # ══════════════════════════════════════════════════════════════
    # EXPORT PDF — ÉTAT DE COMPTE CLIENTS
    # ══════════════════════════════════════════════════════════════
    # ── Fonction helper : construire clients_data depuis les paramètres communs ──
    async def _build_etat_compte_data(
        filtre: str,
        annee_scolaire: Optional[str],
        date_debut: Optional[str],
        date_fin: Optional[str],
        client_id: Optional[str],
        ville: Optional[str],
        type_client: Optional[str],
        representant: Optional[str],
    ):
        """Retourne (clients_data, resume) selon les filtres."""
        query_factures: Dict = {"type_facture": "facture"}

        # Filtre payé / impayé
        if filtre == "paye":
            query_factures["montant_restant"] = 0
        elif filtre == "impaye":
            query_factures["montant_restant"] = {"$gt": 0}

        # Filtre date
        if date_debut or date_fin:
            date_filter: Dict = {}
            if date_debut:
                date_filter["$gte"] = date_debut
            if date_fin:
                date_filter["$lte"] = date_fin
            query_factures["date_facture"] = date_filter

        # Filtre client_id direct
        if client_id:
            query_factures["client_id"] = client_id

        factures_raw = await db.factures.find(
            query_factures,
            {"_id": 0, "facture_id": 1, "reference": 1, "client_id": 1,
             "date_facture": 1, "date_livraison": 1,
             "montant_ht": 1, "montant_ttc": 1, "montant_tva": 1,
             "remise_globale": 1, "montant_regle": 1, "montant_restant": 1,
             "statut": 1}
        ).to_list(None)

        if not factures_raw:
            resume = {
                "nb_clients": 0, "nb_factures": 0,
                "total_vente": 0, "total_remise": 0,
                "total_ht": 0, "total_regle": 0, "total_solde": 0,
            }
            return [], resume

        # Regrouper par client_id
        factures_par_client: Dict[str, List] = defaultdict(list)
        all_client_ids = set()
        for f in factures_raw:
            cid = f.get("client_id")
            if cid:
                factures_par_client[cid].append(f)
                all_client_ids.add(cid)

        # Récupérer clients
        clients_raw = await db.clients.find(
            {"client_id": {"$in": list(all_client_ids)}},
            {"_id": 0, "client_id": 1, "nom": 1, "representant": 1,
             "telephone": 1, "email": 1, "type_client": 1, "region": 1, "ville": 1}
        ).to_list(None)

        # Appliquer filtres client (ville, type_client, representant)
        filtered_clients = []
        for c in clients_raw:
            if ville:
                cv = (c.get("ville") or c.get("region") or "").lower()
                if ville.lower() not in cv:
                    continue
            if type_client and c.get("type_client", "").lower() != type_client.lower():
                continue
            if representant:
                cr = (c.get("representant") or "").lower()
                if representant.lower() not in cr:
                    continue
            filtered_clients.append(c)

        clients_by_id: Dict[str, Dict] = {c["client_id"]: c for c in filtered_clients}

        # Construire clients_data
        clients_data = []
        total_vente  = 0.0
        total_remise = 0.0
        total_ht     = 0.0
        total_regle  = 0.0
        total_solde  = 0.0
        total_facts  = 0

        sorted_client_ids = sorted(
            [cid for cid in factures_par_client.keys() if cid in clients_by_id],
            key=lambda cid: (clients_by_id.get(cid, {}).get("nom") or "").lower()
        )

        for cid in sorted_client_ids:
            client_doc = clients_by_id[cid]
            client_doc["ville"] = (
                client_doc.get("ville") or client_doc.get("region") or "—"
            )
            factures_client = sorted(
                factures_par_client[cid],
                key=lambda x: str(x.get("date_facture") or "")
            )

            clients_data.append({
                "client": client_doc,
                "factures": factures_client,
            })

            for f in factures_client:
                total_vente  += float(f.get("montant_ttc") or 0)
                total_remise += float(f.get("remise_globale") or 0)
                total_ht     += float(f.get("montant_ht") or 0)
                total_regle  += float(f.get("montant_regle") or 0)
                total_solde  += float(f.get("montant_restant") or 0)
                total_facts  += 1

        resume = {
            "nb_clients":   len(clients_data),
            "nb_factures":  total_facts,
            "total_vente":  total_vente,
            "total_remise": total_remise,
            "total_ht":     total_ht,
            "total_regle":  total_regle,
            "total_solde":  total_solde,
        }
        return clients_data, resume

    @router.get("/etat-compte-clients/data")
    async def get_etat_compte_clients_data(
        request: Request,
        filtre: str = "tous",
        annee: Optional[str] = None,
        annee_scolaire: Optional[str] = None,
        date_debut: Optional[str] = None,
        date_fin: Optional[str] = None,
        client_id: Optional[str] = None,
        ville: Optional[str] = None,
        type_client: Optional[str] = None,
        representant: Optional[str] = None,
        authorization: Optional[str] = Header(None),
    ):
        """Retourne les données JSON pour l'aperçu frontend."""
        await check_read(request, authorization)

        clients_data, resume = await _build_etat_compte_data(
            filtre=filtre,
            annee_scolaire=annee_scolaire,
            date_debut=date_debut,
            date_fin=date_fin,
            client_id=client_id,
            ville=ville,
            type_client=type_client,
            representant=representant,
        )
        return {"clients_data": clients_data, "resume": resume}

    @router.get("/etat-compte-clients")
    async def export_etat_compte_clients(
        request: Request,
        filtre: str = "tous",
        annee: Optional[str] = None,
        annee_scolaire: Optional[str] = None,
        date_debut: Optional[str] = None,
        date_fin: Optional[str] = None,
        client_id: Optional[str] = None,
        ville: Optional[str] = None,
        type_client: Optional[str] = None,
        representant: Optional[str] = None,
        authorization: Optional[str] = Header(None),
    ):
        """
        Génère et renvoie le PDF État de Compte Clients.
        filtre: tous | paye | impaye
        annee_scolaire: ex "2024-2025" (affiché dans l'en-tête)
        date_debut / date_fin : filtre sur date_facture (format YYYY-MM-DD)
        client_id, ville, type_client, representant : filtres additionnels
        """
        from fastapi.responses import StreamingResponse
        from compte_client_pdf_generator import generate_etat_compte_clients_pdf

        await check_read(request, authorization)

        now_year = str(datetime.now().year)
        if annee is None:
            annee = now_year

        clients_data, resume = await _build_etat_compte_data(
            filtre=filtre,
            annee_scolaire=annee_scolaire,
            date_debut=date_debut,
            date_fin=date_fin,
            client_id=client_id,
            ville=ville,
            type_client=type_client,
            representant=representant,
        )

        pdf_buf = generate_etat_compte_clients_pdf(
            clients_data, resume, filtre=filtre, annee=annee,
            annee_scolaire=annee_scolaire,
            date_debut=date_debut, date_fin=date_fin,
        )

        filename = f"etat_compte_clients_{filtre}_{annee_scolaire or annee}.pdf"
        return StreamingResponse(
            pdf_buf,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @router.get("/etat-compte-clients/excel")
    async def export_etat_compte_clients_excel(
        request: Request,
        filtre: str = "tous",
        annee: Optional[str] = None,
        annee_scolaire: Optional[str] = None,
        date_debut: Optional[str] = None,
        date_fin: Optional[str] = None,
        client_id: Optional[str] = None,
        ville: Optional[str] = None,
        type_client: Optional[str] = None,
        representant: Optional[str] = None,
        authorization: Optional[str] = Header(None),
    ):
        """Génère et renvoie l'export Excel État de Compte Clients."""
        from fastapi.responses import StreamingResponse
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter

        await check_read(request, authorization)

        now_year = str(datetime.now().year)
        if annee is None:
            annee = now_year

        clients_data, resume = await _build_etat_compte_data(
            filtre=filtre,
            annee_scolaire=annee_scolaire,
            date_debut=date_debut,
            date_fin=date_fin,
            client_id=client_id,
            ville=ville,
            type_client=type_client,
            representant=representant,
        )

        # ── Créer le workbook ──
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "État de Compte Clients"

        # Styles
        BLUE   = "1F4E79"
        ORANGE = "FF6200"
        LBLUE  = "D0E2F7"
        LGREY  = "F3F4F6"
        WHITE  = "FFFFFF"
        RED    = "DC2626"
        GREEN  = "16A34A"

        def hdr_font(bold=True, color=WHITE, size=10):
            return Font(bold=bold, color=color, size=size, name="Calibri")

        def cell_font(bold=False, color="0A2540", size=9):
            return Font(bold=bold, color=color, size=size, name="Calibri")

        def fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def thin_border():
            s = Side(style="thin", color="D1D5DB")
            return Border(left=s, right=s, top=s, bottom=s)

        def center():
            return Alignment(horizontal="center", vertical="center", wrap_text=True)

        def right_align():
            return Alignment(horizontal="right", vertical="center")

        # ── Titre global ──
        ws.merge_cells("A1:H1")
        c = ws["A1"]
        c.value = "ÉDITIONS FABS-CI — ÉTAT DE COMPTE CLIENTS"
        c.font = Font(bold=True, color=WHITE, size=13, name="Calibri")
        c.fill = fill(BLUE)
        c.alignment = center()
        ws.row_dimensions[1].height = 22

        # Sous-titre
        periode = ""
        if annee_scolaire:
            periode = f"Année scolaire : {annee_scolaire}"
        elif date_debut or date_fin:
            periode = f"Période : {date_debut or '…'} → {date_fin or '…'}"
        else:
            periode = f"Année : {annee}"

        ws.merge_cells("A2:H2")
        c2 = ws["A2"]
        c2.value = f"Filtre : {filtre.upper()}   |   {periode}   |   Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        c2.font = Font(bold=False, color=ORANGE, size=9, name="Calibri")
        c2.fill = fill("EFF6FF")
        c2.alignment = center()
        ws.row_dimensions[2].height = 16

        ws.append([])  # ligne vide
        ws.row_dimensions[3].height = 6

        # ── Colonnes ──
        COLS = ["Référence", "Date facture", "Montant TTC", "Remise", "Montant HT", "Réglé", "Solde restant", "Statut"]
        COL_WIDTHS = [18, 14, 16, 12, 16, 16, 16, 14]
        for i, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        row_idx = 4
        fmt_money = '#,##0.00 "FCFA"'
        FILTRE_LABELS = {"tous": "Tous", "paye": "Payés", "impaye": "Impayés"}

        for entry in clients_data:
            client = entry["client"]
            factures = entry["factures"]

            # ── Ligne client ──
            ws.merge_cells(f"A{row_idx}:H{row_idx}")
            nom_cell = ws.cell(row_idx, 1)
            infos = f"  {client.get('nom', '—')}   |   Ville : {client.get('ville','—')}   |   Type : {client.get('type_client','—')}   |   Représentant : {client.get('representant','—')}"
            nom_cell.value = infos
            nom_cell.font = Font(bold=True, color=WHITE, size=9.5, name="Calibri")
            nom_cell.fill = fill(BLUE)
            nom_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row_idx].height = 16
            row_idx += 1

            # En-tête colonnes
            for ci, col_name in enumerate(COLS, 1):
                hc = ws.cell(row_idx, ci)
                hc.value = col_name
                hc.font = hdr_font(color=WHITE, size=8.5)
                hc.fill = fill("2563EB")
                hc.alignment = center()
                hc.border = thin_border()
            ws.row_dimensions[row_idx].height = 14
            row_idx += 1

            # Lignes factures
            sub_ttc = sub_remise = sub_ht = sub_regle = sub_solde = 0.0
            for fi, f in enumerate(factures):
                ttc    = float(f.get("montant_ttc") or 0)
                remise = float(f.get("remise_globale") or 0)
                ht     = float(f.get("montant_ht") or 0)
                regle  = float(f.get("montant_regle") or 0)
                solde  = float(f.get("montant_restant") or 0)
                statut = f.get("statut", "—")
                sub_ttc += ttc; sub_remise += remise; sub_ht += ht
                sub_regle += regle; sub_solde += solde

                row_fill = fill(WHITE) if fi % 2 == 0 else fill(LGREY)
                vals = [f.get("reference","—"), f.get("date_facture","—"),
                        ttc, remise, ht, regle, solde, statut]
                for ci, v in enumerate(vals, 1):
                    dc = ws.cell(row_idx, ci)
                    dc.value = v
                    dc.font = cell_font()
                    dc.fill = row_fill
                    dc.border = thin_border()
                    if ci in (3, 4, 5, 6, 7):
                        dc.number_format = fmt_money
                        dc.alignment = right_align()
                    elif ci == 2:
                        dc.alignment = center()
                    else:
                        dc.alignment = Alignment(horizontal="left", vertical="center")
                    # Couleur solde
                    if ci == 7:
                        dc.font = Font(bold=True, color=(RED if solde > 0 else GREEN), size=9, name="Calibri")
                ws.row_dimensions[row_idx].height = 13
                row_idx += 1

            # Sous-total client
            ws.merge_cells(f"A{row_idx}:B{row_idx}")
            stc = ws.cell(row_idx, 1)
            stc.value = f"SOUS-TOTAL — {client.get('nom','')}"
            stc.font = Font(bold=True, color=BLUE, size=8.5, name="Calibri")
            stc.fill = fill(LBLUE)
            stc.alignment = Alignment(horizontal="left", vertical="center")
            for ci, v in [(3, sub_ttc),(4, sub_remise),(5, sub_ht),(6, sub_regle),(7, sub_solde)]:
                sc = ws.cell(row_idx, ci)
                sc.value = v
                sc.font = Font(bold=True, color=BLUE, size=8.5, name="Calibri")
                sc.fill = fill(LBLUE)
                sc.number_format = fmt_money
                sc.alignment = right_align()
                sc.border = thin_border()
            ws.cell(row_idx, 8).fill = fill(LBLUE)
            ws.row_dimensions[row_idx].height = 14
            row_idx += 2  # saut de ligne entre clients

        # ── Total général ──
        ws.merge_cells(f"A{row_idx}:B{row_idx}")
        tc = ws.cell(row_idx, 1)
        tc.value = f"TOTAL GÉNÉRAL ({resume['nb_clients']} clients — {resume['nb_factures']} factures)"
        tc.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
        tc.fill = fill(ORANGE)
        tc.alignment = Alignment(horizontal="left", vertical="center")
        for ci, v in [
            (3, resume["total_vente"]), (4, resume["total_remise"]),
            (5, resume["total_ht"]),    (6, resume["total_regle"]),
            (7, resume["total_solde"]),
        ]:
            gc = ws.cell(row_idx, ci)
            gc.value = v
            gc.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
            gc.fill = fill(ORANGE)
            gc.number_format = fmt_money
            gc.alignment = right_align()
        ws.cell(row_idx, 8).fill = fill(ORANGE)
        ws.row_dimensions[row_idx].height = 18

        # Figer la ligne de titre
        ws.freeze_panes = "A4"

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_periode = (annee_scolaire or annee).replace("/", "-")
        filename = f"etat_compte_clients_{filtre}_{safe_periode}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    return router
